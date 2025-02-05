#!/usr/bin/env python3

import openpyxl
import os.path

#Create a workbook
if os.path.isfile('strg/otro.xlsx'):
    print('EL archivo ya existe no se sobreescribirá.')
else:
    ssBook = openpyxl.Workbook()
    ssBook.save('strg/otro.xlsx')
    print('Archivo creado con éxito.')
    ssBook.close()

wb = openpyxl.load_workbook(filename='strg/otro.xlsx')
print(wb.sheetnames)
