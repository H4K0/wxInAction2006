#!/usr/bin/env python3

import openpyxl
import os.path

#Create a workbook
if os.path.isfile('strg/9° A.xlsx'):
    print('El archivo ya existe no se sobreescribirá.')
    pass
else:
    print('El archivo no debe ser creado.')
    exit()

wb = openpyxl.load_workbook(filename='strg/9° A.xlsx')
print(wb.sheetnames)
sheet = wb.active
print('Sheet:', sheet)

print('Sheet title:', sheet.title)

wb.create_sheet('Salida')
print(wb.sheetnames)

wb.active[1]
print(wb.active)

wb.active["A1"] = "H4K0 PROOF ;)"

wb.save(filename='strg/9° A.xlsx')

wb.close