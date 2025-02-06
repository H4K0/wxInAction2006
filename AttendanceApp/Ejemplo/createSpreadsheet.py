#!/usr/bin/env python3

import openpyxl

wb = openpyxl.load_workbook(filename='strg/ssFile.xlsx')
#ws0 = ssBook.active
print(wb.sheetnames)
print(len(wb.sheetnames))

wb.active = 0
print(wb.active)


data = [
    ('Nombre', 'Edad', 'Compañía'),
    ('John', 21, 'Facebook'),
    ('Hannah', 21, 'Apple'),
    ('Camila', 27, 'Toyota'),
    ('Steve', 32, 'Google')
]

for i, j in enumerate(data):
    wb.active[f'A{i+1}'] = j[0]
    wb.active[f'B{i+1}'] = j[1]
    wb.active[f'c{i+1}'] = j[2]


wb.active = 1
print(wb.active)

wb.active["A1"] = "writing ;)"

wb.save(filename='strg/ssFile.xlsx')
wb.close