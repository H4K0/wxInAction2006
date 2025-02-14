import openpyxl

#Crea un archivo, si existe, lo sobrescribe.
"""
wb = openpyxl.Workbook()
wb.save('AttendanceApp/Ejemplo/strg/otro2.xlsx')

ws1 = wb.create_sheet('Hoja1')
ws2 = wb.create_sheet('Hoja2', 0)
ws3 = wb.create_sheet('Hoja3',-1)

ws1.title = 'Hoja1Nueva'
ws2.title = 'Hoja2Nueva'

ws3.sheet_properties.tabColor = '1072BA'

wb.save('AttendanceApp/Ejemplo/strg/otro2.xlsx')
"""


data = [
    ('Nombre', 'Edad', 'Compañía'),
    ('John', 21, 'Facebook'),
    ('Hannah', 21, 'Apple'),
    ('Camila', 27, 'Toyota'),
    ('Steve', 32, 'Google')
]

wb = openpyxl.load_workbook('AttendanceApp/Ejemplo/strg/otro2.xlsx')

#newSheet = wb.create_sheet('Nueva')

print(wb.sheetnames)

listNames = wb.sheetnames

print(listNames)
print(type(listNames))
activeSheet = wb.active

wb.active = wb['Nueva']
wb.active["A1"] = "H4K0 PROOF ;)"
print('Active: ', wb.active)

wb.active = wb['Sheet']
wb.active["A1"] = "H4K0 PROOF ;)"
print('Active: ', wb.active)



"""
for h in range(len(listNames)):
    wb.active = h
    print('Active:', h, wb.active)
    for i, j in enumerate(data):
        wb.active[f'X{i+1}'] = j[0]
        wb.active[f'Y{i+1}'] = j[1]
        wb.active[f'Z{i+1}'] = j[2]
    print('Hecho!')
"""


"""
wb.active = wb['Datos']
print('Active:', wb.active)
wb.active = 7
print('Active:', wb.active)
"""


datos = wb['Datos']

datos['D3'] = 'Octavio'

x = datos['D5']
print(datos['D5'].value)
print(x.value)

if x.value == 'Abraxas':
    print('Muy bien sí es.')
else:
    print('Fallaste')



wb.save('AttendanceApp/Ejemplo/strg/otro2.xlsx')
wb.close